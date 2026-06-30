#!/usr/bin/env python3
"""
Analyze datamigration.xlsx - list sheets, headers, sample data
"""
import openpyxl
from openpyxl.utils import get_column_letter

FILE = '/Users/a1/Documents/motivation/fundamental-engine/datamigration.xlsx'
MAX_SAMPLE_ROWS = 5
MAX_COLS = 30

def cell_val(cell):
    if cell is None:
        return ''
    v = cell.value
    if v is None:
        return ''
    return str(v)[:80]

wb = openpyxl.load_workbook(FILE, read_only=True, data_only=True)
sheets = wb.sheetnames
print(f'=== TOTAL SHEETS: {len(sheets)} ===\n')
for i, name in enumerate(sheets):
    print(f'{i+1:2}. [{name}]')

print('\n' + '='*80)

for sheet_name in sheets:
    ws = wb[sheet_name]
    print(f'\n{"="*80}')
    print(f'SHEET: [{sheet_name}]')
    print(f'{"="*80}')
    
    rows_read = 0
    for row in ws.iter_rows(max_row=MAX_SAMPLE_ROWS + 3, max_col=MAX_COLS):
        if rows_read >= MAX_SAMPLE_ROWS:
            break
        row_vals = [cell_val(c) for c in row]
        # skip entirely empty rows
        if not any(v.strip() for v in row_vals):
            continue
        # trim trailing empty
        while row_vals and not row_vals[-1].strip():
            row_vals.pop()
        print(' | '.join(f'{v:<20}' for v in row_vals))
        rows_read += 1
    
    print(f'  (max_row≈{ws.max_row}, max_col≈{ws.max_column})')

wb.close()
print('\n=== DONE ===')
